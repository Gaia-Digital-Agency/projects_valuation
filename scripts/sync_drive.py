#!/usr/bin/env python3
"""
Sync a single Google Drive folder (recursively) into the valuations.dd_files table.
Usage: python3 sync_drive.py <company_id>
Reads the company's drive_folder_id from Postgres.
"""
import os, sys, io, json
from datetime import datetime, timezone
from pathlib import Path

from dotenv import dotenv_values
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import psycopg2
from psycopg2.extras import execute_values

ENV = dotenv_values(Path(__file__).resolve().parent.parent / 'backend' / '.env')

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
CREDS = os.path.expanduser(ENV.get('GOOGLE_CREDENTIALS_PATH', '~/.credentials/drive_credentials.json'))
TOKEN = os.path.expanduser(ENV.get('GOOGLE_TOKEN_PATH', '~/.credentials/drive_token.json'))

# Extraction
EXTRACT_MIME = {
    'application/pdf',
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    'application/vnd.google-apps.document',
    'application/vnd.google-apps.spreadsheet',
    'text/plain', 'text/csv',
}
MAX_TEXT = 200_000  # cap per file


def auth():
    creds = None
    if os.path.exists(TOKEN):
        creds = Credentials.from_authorized_user_file(TOKEN, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDS, SCOPES)
            creds = flow.run_local_server(port=0)
        Path(TOKEN).write_text(creds.to_json())
    return build('drive', 'v3', credentials=creds)


def list_folder(svc, folder_id, path=''):
    """Recursively list all files under folder_id. Yields metadata dicts."""
    page = None
    while True:
        resp = svc.files().list(
            q=f"'{folder_id}' in parents and trashed=false",
            fields='nextPageToken, files(id, name, mimeType, size, modifiedTime, webViewLink, parents)',
            pageSize=200, pageToken=page, supportsAllDrives=True, includeItemsFromAllDrives=True,
        ).execute()
        for f in resp.get('files', []):
            if f['mimeType'] == 'application/vnd.google-apps.folder':
                sub = f"{path}/{f['name']}" if path else f['name']
                yield from list_folder(svc, f['id'], sub)
            else:
                f['_folder_path'] = path or '/'
                yield f
        page = resp.get('nextPageToken')
        if not page:
            break


def download(svc, file_id, mime):
    """Download or export file as bytes. Returns (bytes, effective_mime) or (None, mime)."""
    try:
        if mime == 'application/vnd.google-apps.document':
            req = svc.files().export_media(fileId=file_id,
                mimeType='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
            mime = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        elif mime == 'application/vnd.google-apps.spreadsheet':
            req = svc.files().export_media(fileId=file_id,
                mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            req = svc.files().get_media(fileId=file_id, supportsAllDrives=True)
        buf = io.BytesIO()
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        return buf.getvalue(), mime
    except Exception as e:
        print(f"  download failed: {e}", file=sys.stderr)
        return None, mime


def extract_text(data, mime, name):
    if not data:
        return None
    try:
        if 'pdf' in mime:
            from pypdf import PdfReader
            r = PdfReader(io.BytesIO(data))
            return '\n'.join((p.extract_text() or '') for p in r.pages)[:MAX_TEXT]
        if 'spreadsheetml' in mime or name.lower().endswith('.xlsx'):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            chunks = []
            for ws in wb.worksheets:
                chunks.append(f'\n=== Sheet: {ws.title} ===')
                for row in ws.iter_rows(values_only=True):
                    chunks.append('\t'.join('' if v is None else str(v) for v in row))
                if sum(len(c) for c in chunks) > MAX_TEXT:
                    break
            return '\n'.join(chunks)[:MAX_TEXT]
        if 'wordprocessingml' in mime or name.lower().endswith('.docx'):
            from docx import Document
            d = Document(io.BytesIO(data))
            return '\n'.join(p.text for p in d.paragraphs)[:MAX_TEXT]
        if mime.startswith('text/'):
            return data.decode('utf-8', errors='replace')[:MAX_TEXT]
    except Exception as e:
        print(f"  extract failed: {e}", file=sys.stderr)
    return None


def main():
    company_id = sys.argv[1] if len(sys.argv) > 1 else 'greenviro'
    conn = psycopg2.connect(
        host=ENV['PGHOST'], port=ENV['PGPORT'], user=ENV['PGUSER'],
        password=ENV['PGPASSWORD'], dbname=ENV['PGDATABASE'])
    conn.autocommit = False
    cur = conn.cursor()

    cur.execute("SELECT drive_folder_id FROM companies WHERE id = %s", (company_id,))
    row = cur.fetchone()
    if not row:
        print(f"Unknown company: {company_id}", file=sys.stderr); sys.exit(1)
    folder_id = row[0]
    print(f"[{company_id}] folder={folder_id}")

    svc = auth()
    files = list(list_folder(svc, folder_id))
    print(f"  found {len(files)} files")

    rows = []
    for i, f in enumerate(files, 1):
        text = None
        if f['mimeType'] in EXTRACT_MIME:
            print(f"  [{i}/{len(files)}] {f['name']}")
            data, eff_mime = download(svc, f['id'], f['mimeType'])
            text = extract_text(data, eff_mime, f['name'])
            if text:
                text = text.replace('\x00', '')
        else:
            print(f"  [{i}/{len(files)}] {f['name']} (skip extract: {f['mimeType']})")
        rows.append((
            company_id, f['id'], f['name'], f['mimeType'],
            int(f.get('size') or 0) or None,
            f['_folder_path'], f.get('webViewLink'),
            f.get('modifiedTime'), text,
            datetime.now(timezone.utc) if text else None,
        ))

    execute_values(cur, """
        INSERT INTO dd_files (company_id, drive_id, name, mime_type, size_bytes,
                              folder_path, web_view_link, modified_time, extracted_text, extracted_at)
        VALUES %s
        ON CONFLICT (company_id, drive_id) DO UPDATE SET
            name = EXCLUDED.name,
            mime_type = EXCLUDED.mime_type,
            size_bytes = EXCLUDED.size_bytes,
            folder_path = EXCLUDED.folder_path,
            web_view_link = EXCLUDED.web_view_link,
            modified_time = EXCLUDED.modified_time,
            extracted_text = COALESCE(EXCLUDED.extracted_text, dd_files.extracted_text),
            extracted_at = COALESCE(EXCLUDED.extracted_at, dd_files.extracted_at),
            indexed_at = NOW()
    """, rows)
    conn.commit()
    print(f"imported={len(rows)}")
    print(json.dumps({'imported': len(rows)}))


if __name__ == '__main__':
    main()
